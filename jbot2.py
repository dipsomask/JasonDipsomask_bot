import telebot
import pandas as pd
from telebot import types
import os
import json
import shutil
import time
from openpyxl import load_workbook

import pathes
import botToken

# static parms

jd = telebot.TeleBot(botToken.TOKEN)

mode = True # user - True / manager - False

if(mode):
    admin = botToken.admin1
    admin_lnk = botToken.admin_lnk1
    admin_chat_id = botToken.admin_chat_id1 
else:
    admin = botToken.admin2
    admin_lnk = botToken.admin_lnk2
    admin_chat_id = botToken.admin_chat_id2


with open(pathes.mainfile_copy_json) as jsf:
    allmaindata = json.load(jsf)


last_firm = ""

btnToMain = types.KeyboardButton("Главная")

new_user = {
    "chatid": "",
    "username": ""
}

#


def saveChatId(chat_id, user):
    with open(pathes.all_users_chatid, 'r', encoding='utf-8') as file:
        save_chats = json.load(file)
    chats = []
    if save_chats:
        for blok in save_chats:
            chats.append(blok["chatid"])
    if not (chat_id in chats):
        new_user["chatid"] = chat_id
        new_user["username"] = user
        save_chats.append(new_user)
        with open(pathes.all_users_chatid, 'w', encoding='utf-8') as file:
                json.dump(save_chats, file, ensure_ascii=False, indent=4)
        return f"chat_id {chat_id}, добавлен."
    else:
        return f"chat_id {chat_id}, уже существует."


def parseMainDataJson(position):
    return allmaindata[position]


def parseBasket(user):
    filepath = pathes.baskets_user_basket_json + str(user) + '.json'
    if(not os.path.exists(filepath)):
        return "Корзина пуста"
    string_data = ""

    with open(filepath, 'r', encoding='utf-8') as file:
        data = json.load(file)

    for firm, products in data.items():
        if(products):
            for product in products:
                string_data += f"{product["Наименование товара"]}, {product["Цена_наша"]}р, {product["Кол-во"]}шт\n"

    return string_data


def makeXLSXuserBasket(user):
    filepath = pathes.baskets_user_basket_json + str(user) + '.json'
    xlsx_filepath = pathes.baskets_in_agreement_xlsx + str(user) + '.xlsx'
    json_filepath = pathes.baskets_in_agreement_json + str(user) + '.json'
    with open(filepath, 'r', encoding='utf-8') as file:
        data = json.load(file)

    filtered_data = []
    control_summ = 0

    for key, items in data.items():
        for item in items:
            if 'Цена_наша' in item:
                control_summ += item['Сумма'] 
            filtered_item = {k: v for k, v in item.items() if k != 'Цена'}
            filtered_data.append(filtered_item)

    last_str = {
            "Номер":"",
            "Наименование товара": "",
            "Цена_наша":"",
            "Кол-во": "",
            "Ед. изм.": "И того",
            "Сумма": control_summ
        }
    
    filtered_data.append(last_str)

    df = pd.DataFrame(filtered_data)
    df.to_excel(xlsx_filepath, index=False)

    workbook = load_workbook(xlsx_filepath)
    sheet = workbook.active

    sheet.column_dimensions['B'].width = 42.67 

    workbook.save(xlsx_filepath)

    shutil.move(filepath, json_filepath)

    return xlsx_filepath


def makeXLSXourBasket():
    json_filepath = pathes.our_basket_json
    if not os.path.exists(json_filepath):
        return "Нет заказов"
    else:
        with open(json_filepath, 'r', encoding='utf-8') as file:
            data = json.load(file)

        items = []
        for key in data:
            items.extend(data[key])

        control_summ = 0

        filtered_items = []
        for item in items:
            if 'Сумма' in item:
                control_summ += item['Сумма']
            filtered_item = {k: v for k, v in item.items() if k != 'Цена_наша'}
            filtered_items.append(filtered_item)

        last_str = {
            "Наименование товара": "",
            "Цена": "",
            "Кол-во": "",
            "Ед. изм.": "Итого",
            "Сумма": control_summ
        }

        filtered_items.append(last_str)

        xlsx_filepath = pathes.our_basket_xlsx
        df = pd.DataFrame(filtered_items)
        df.to_excel(xlsx_filepath, index=False)

        workbook = load_workbook(xlsx_filepath)
        sheet = workbook.active

        sheet.column_dimensions['B'].width = 42.67 

        workbook.save(xlsx_filepath)

        return xlsx_filepath


def addProductToUserBaasket(user, firm, product):
    filepath = pathes.baskets_user_basket_json + str(user) + '.json'
    if(not os.path.exists(filepath)):
        source_file = pathes.user_basket_json
        shutil.copy(source_file, filepath)

    for blok in allmaindata[firm]:
        if(blok["Наименование товара"] == product):
            newblokinbasket = blok
            break

    with open(filepath, 'r', encoding='utf-8') as file:
        data = json.load(file)
    
    thereis = False

    for blok in data[firm]:
        if(blok["Наименование товара"] == newblokinbasket["Наименование товара"]):
            blok["Кол-во"] += 1
            blok["Сумма"] = blok["Кол-во"] * blok["Цена_наша"]
            thereis = True
            break
    
    if(thereis is False):
        data[firm].append(newblokinbasket)

    with open(filepath, 'w', encoding='utf-8') as file:
        json.dump(data, file, ensure_ascii=False, indent=4)

    return 'Операция завершена'


def parseAgreements():
    dirpath = pathes.agreement_dir_xlsx
    files = []
    for root, dirs, filenames in os.walk(dirpath):
        files.extend(filenames)

    return files


def aplyOrNot(user, mode):
    sourse_json = pathes.baskets_in_agreement_json + user + '.json'
    sourse_xlsx = pathes.baskets_in_agreement_xlsx + user + '.xlsx'
    if(not mode):
        if(os.path.exists(sourse_json)):
            os.remove(sourse_json)
        if(os.path.exists(sourse_xlsx)):
            os.remove(sourse_xlsx)
        return "Заявка отклонена."
    else:
        ourbasket_json = pathes.our_basket_json
        if(not os.path.exists(ourbasket_json)):
            source_file = pathes.our_basket_empty_json
            shutil.copy(source_file, ourbasket_json)
        user_agreement = pathes.baskets_in_agreement_json + user + '.json'

        with open(user_agreement, 'r', encoding='utf-8') as file:
            user_agreement_data = json.load(file)


        with open(ourbasket_json, 'r', encoding='utf-8') as file:
            ourbasket_json_data = json.load(file)

        for key in user_agreement_data:
            if user_agreement_data[key]:
                for user_product in user_agreement_data[key]:
                    for our_product in ourbasket_json_data[key]:
                        if(user_product["Наименование товара"] == our_product["Наименование товара"]):
                            our_product["Кол-во"] += user_product["Кол-во"]
                            our_product["Сумма"] = our_product["Кол-во"] * our_product["Цена"]

        with open(ourbasket_json, 'w', encoding='utf-8') as json_file:
            json.dump(ourbasket_json_data, json_file, ensure_ascii=False, indent=4)

        dest_json = pathes.archive_user_basket_json + user + '.json'
        dest_xlsx = pathes.archive_user_basket_xlsx + user + '.xlsx'
        shutil.move(sourse_json, dest_json)
        shutil.move(sourse_xlsx, dest_xlsx)

        return "Заявка одобрена"
        



@jd.message_handler(commands=['start'])
def start(message):
    user = message.from_user.username
    chat_id = message.chat.id

    print(saveChatId(chat_id, user))

    print(f"user={user}, id={chat_id}, command={message.text}, mode={mode}")
    if(user != admin):
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        btn1 = types.KeyboardButton("Корзина")
        btn2 = types.KeyboardButton("Каталог")
        markup.add(btn1, btn2)
        jd.send_message(message.chat.id, text="Привет, {0.first_name}! Я бот для JDShop".format(message.from_user), reply_markup=markup)
    else:
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        btn1 = types.KeyboardButton("Показать заявки")
        btn2 = types.KeyboardButton("Скачать общую кокрзину")
        btn3 = types.KeyboardButton("Создать оповещение")
        markup.add(btn1, btn2, btn3)
        jd.send_message(message.chat.id, text="Привет, {0.first_name}! Я бот для JDShop, а ты Админ.".format(message.from_user), reply_markup=markup)
        

@jd.message_handler(content_types=['text'])
def func(message):
    user = message.from_user.username
    chat_id = message.chat.id
    global last_firm

    print(saveChatId(chat_id, user))

    print(f"user={user}, id={chat_id}, command={message.text}, mode={mode}, last_firm={last_firm}")
    if(user != admin):
        if(message.text == "Каталог"):
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            btn1 = types.KeyboardButton("ZoNk!")
            btn2 = types.KeyboardButton("Odb Juice")
            btn3 = types.KeyboardButton("Juice Head")
            btn4 = types.KeyboardButton("Juice Man")
            btn5 = types.KeyboardButton("Custard Monster")
            btn6 = types.KeyboardButton("FRUIT MONSTER")
            btn7 = types.KeyboardButton("Jam Monster")
            btn8 = types.KeyboardButton("Lemonade Monster")
            btn9 = types.KeyboardButton("Milk Paradise")
            btn10 = types.KeyboardButton("Bakery Vapor")
            btn11 = types.KeyboardButton("Sweet Collection")
            markup.add(btn1, btn2, btn3, btn4, btn5, btn6, btn7, btn8, btn9, btn10, btn11)
            jd.send_message(message.chat.id, text="Выберите производителя:", reply_markup=markup)

        elif(message.text == "ZoNk!"):
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            data = parseMainDataJson("ZoNk!")
            last_firm = "ZoNk!"
            for blok in data:
                markup.add(types.KeyboardButton("Добавить в корзину: " + blok["Наименование товара"] + " " + str(blok["Цена_наша"]) + "р"))
            markup.add(btnToMain)
            jd.send_message(message.chat.id, text="Нажмите на товар для добавления в корзину.", reply_markup=markup)
        elif(message.text == "Odb Juice"):
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            data = parseMainDataJson("Odb Juice")
            last_firm = "Odb Juice"
            for blok in data:
                markup.add(types.KeyboardButton("Добавить в корзину: " + blok["Наименование товара"] + " " + str(blok["Цена_наша"]) + "р"))
            markup.add(btnToMain)
            jd.send_message(message.chat.id, text="Нажмите на товар для добавления в корзину.", reply_markup=markup)
        elif(message.text == "Juice Head"):
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            data = parseMainDataJson("Juice Head")
            last_firm = "Juice Head"
            for blok in data:
                markup.add(types.KeyboardButton("Добавить в корзину: " + blok["Наименование товара"] + " " + str(blok["Цена_наша"]) + "р"))
            markup.add(btnToMain)
            jd.send_message(message.chat.id, text="Нажмите на товар для добавления в корзину.", reply_markup=markup)
        elif(message.text == "Juice Man"):
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            data = parseMainDataJson("Juice Man")
            last_firm = "Juice Man"
            for blok in data:
                markup.add(types.KeyboardButton("Добавить в корзину: " + blok["Наименование товара"] + " " + str(blok["Цена_наша"]) + "р"))
            markup.add(btnToMain)
            jd.send_message(message.chat.id, text="Нажмите на товар для добавления в корзину.", reply_markup=markup)
        elif(message.text == "Custard Monster"):
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            data = parseMainDataJson("Custard Monster")
            last_firm = "Custard Monster"
            for blok in data:
                markup.add(types.KeyboardButton("Добавить в корзину: " + blok["Наименование товара"] + " " + str(blok["Цена_наша"]) + "р"))
            markup.add(btnToMain)
            jd.send_message(message.chat.id, text="Нажмите на товар для добавления в корзину.", reply_markup=markup)
        elif(message.text == "FRUIT MONSTER"):
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            data = parseMainDataJson("FRUIT MONSTER")
            last_firm = "FRUIT MONSTER"
            for blok in data:
                markup.add(types.KeyboardButton("Добавить в корзину: " + blok["Наименование товара"] + " " + str(blok["Цена_наша"]) + "р"))
            markup.add(btnToMain)
            jd.send_message(message.chat.id, text="Нажмите на товар для добавления в корзину.", reply_markup=markup)
        elif(message.text == "Jam Monster"):
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            data = parseMainDataJson("Jam Monster")
            last_firm = "Jam Monster"
            for blok in data:
                markup.add(types.KeyboardButton("Добавить в корзину: " + blok["Наименование товара"] + " " + str(blok["Цена_наша"]) + "р"))
            markup.add(btnToMain)
            jd.send_message(message.chat.id, text="Нажмите на товар для добавления в корзину.", reply_markup=markup)
        elif(message.text == "Lemonade Monster"):
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            data = parseMainDataJson("Lemonade Monster")
            last_firm = "Lemonade Monster"
            for blok in data:
                markup.add(types.KeyboardButton("Добавить в корзину: " + blok["Наименование товара"] + " " + str(blok["Цена_наша"]) + "р"))
            markup.add(btnToMain)
            jd.send_message(message.chat.id, text="Нажмите на товар для добавления в корзину.", reply_markup=markup)
        elif(message.text == "Milk Paradise"):
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            data = parseMainDataJson("Milk Paradise")
            last_firm = "Milk Paradise"
            for blok in data:
                markup.add(types.KeyboardButton("Добавить в корзину: " + blok["Наименование товара"] + " " + str(blok["Цена_наша"]) + "р"))
            markup.add(btnToMain)
            jd.send_message(message.chat.id, text="Нажмите на товар для добавления в корзину.", reply_markup=markup)
        elif(message.text == "Bakery Vapor"):
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            data = parseMainDataJson("Bakery Vapor")
            last_firm = "Bakery Vapor"
            for blok in data:
                markup.add(types.KeyboardButton("Добавить в корзину: " + blok["Наименование товара"] + " " + str(blok["Цена_наша"]) + "р"))
            markup.add(btnToMain)
            jd.send_message(message.chat.id, text="Нажмите на товар для добавления в корзину.", reply_markup=markup)
        elif(message.text == "Sweet Collection"):
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            data = parseMainDataJson("Sweet Collection")
            last_firm = "Sweet Collection"
            for blok in data:
                markup.add(types.KeyboardButton("Добавить в корзину: " + blok["Наименование товара"] + " " + str(blok["Цена_наша"]) + "р"))
            markup.add(btnToMain)
            jd.send_message(message.chat.id, text="Нажмите на товар для добавления в корзину.", reply_markup=markup)


        elif(message.text == "Главная"):
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            btn1 = types.KeyboardButton("Корзина")
            btn2 = types.KeyboardButton("Каталог")
            markup.add(btn1, btn2)
            jd.send_message(message.chat.id, text="Вы на главной.", reply_markup=markup)


        elif message.text.startswith("Добавить в корзину: "):
            product = message.text.replace("Добавить в корзину: ", "")
            last_space_index = product.rfind(' ')
            if last_space_index != -1:
                result = product[:last_space_index]
            else:
                result = product
            
            jd.send_message(message.chat.id, text=addProductToUserBaasket(user, last_firm, result))


        elif(message.text == "Корзина"):
            answer = parseBasket(user)
            if(answer != "Корзина пуста"):
                markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
                btn1 = types.KeyboardButton("Создать заявку")
                btn2 = types.KeyboardButton("Очистить корзину")
                markup.add(btn1, btn2, btnToMain)
                jd.send_message(message.chat.id, text=answer, reply_markup=markup)
            else:
                jd.send_message(message.chat.id, text=answer)

        elif(message.text == "Очистить корзину"):
            filepath = filepath = pathes.baskets_user_basket_json + str(user) + '.json'
            if(not os.path.exists(filepath)):
                jd.send_message(message.chat.id, text="Корзина пуста")
            else:
                os.remove(filepath)
                jd.send_message(message.chat.id, text="Корзина очищена")
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            btn1 = types.KeyboardButton("Корзина")
            btn2 = types.KeyboardButton("Каталог")
            markup.add(btn1, btn2)
            jd.send_message(message.chat.id, text="Вы на главной странице", reply_markup=markup)
        

        elif(message.text == "Создать заявку"):
            filepath = pathes.baskets_in_agreement_json + user + '.json'
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            btn1 = types.KeyboardButton("Корзина")
            btn2 = types.KeyboardButton("Каталог")
            markup.add(btn1, btn2)
            if(not os.path.exists(filepath)):
                jd.send_message(message.chat.id, text="Заявка отправлена на согласование\В течении дня с вами свяжется менаджер для согласования заявки.")
                with open(makeXLSXuserBasket(user), 'rb') as document:
                    caption_text = 'Клиент ждёт согласования: ' + 'https://t.me/' + user
                    jd.send_document(admin_chat_id, document, caption=caption_text)
            else:
                jd.send_message(message.chat.id, text="Дождитесь окончания обработки предыдущей заявки.")
            jd.send_message(message.chat.id, text="Вы на главной странице", reply_markup=markup)
        
        else:
            jd.send_message(message.chat.id, text="Такая команда не обрабатывается.")

    else:
        if(message.text == "Показать заявки"):
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            filepath = '/home/dipsomask/Документы/JasonDipsomask_bot/jdvenv/baskets-in-agreement/xlsx/'
            agrimenst = parseAgreements()
            if agrimenst:
                for filename in agrimenst:
                    markup.add(types.KeyboardButton(f"Обработать заявку {filename}"))
            
                jd.send_message(message.chat.id, text="Вот список всех заявок.", reply_markup=markup)
            else:
                jd.send_message(message.chat.id, text="Заявки отсутствуют", reply_markup=markup)
        
        elif(message.text == "Главная"):
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            btn1 = types.KeyboardButton("Показать заявки")
            btn2 = types.KeyboardButton("Скачать общую кокрзину")
            btn3 = types.KeyboardButton("Создать оповещение")
            markup.add(btn1, btn2, btn3)
            jd.send_message(message.chat.id, text="Вы на главной.", reply_markup=markup)
        
        elif(message.text.startswith("Обработать заявку ")):
            filename = message.text.replace("Обработать заявку ", "")
            filepath = '/home/dipsomask/Документы/JasonDipsomask_bot/jdvenv/baskets-in-agreement/xlsx/' + filename
            agreementuser = (filename.replace(".xlsx", "")).replace("user-basket-", "")
            with open(filepath, 'rb') as document:
                caption_text = 'Заявка от клиента: ' + 'https://t.me/' + agreementuser
                jd.send_document(message.chat.id, document, caption=caption_text)
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            btn1 = types.KeyboardButton(f"Одобрить заявку {filename}")
            btn2 = types.KeyboardButton(f"Отклонить заявку {filename}")
            markup.add(btn1, btn2, btnToMain)
            jd.send_message(message.chat.id, text="Одобрить заявку или нет?", reply_markup=markup)

        elif(message.text.startswith("Одобрить заявку ")):
            filename = message.text.replace("Одобрить заявку ", "")
            filepath = '/home/dipsomask/Документы/JasonDipsomask_bot/jdvenv/baskets-in-agreement/xlsx/' + filename
            agreementuser = (filename.replace(".xlsx", "")).replace("user-basket-", "")
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            markup.add(btnToMain)
            aplyOrNot(agreementuser, True)
            jd.send_message(message.chat.id, text=f"Заявка {filename}, клиента https://t.me/{agreementuser} " +
                            "одобрена и отправлена в общую корзину.", reply_markup=markup)
            
        elif(message.text.startswith("Отклонить заявку ")):
            filename = message.text.replace("Отклонить заявку ", "")
            filepath = '/home/dipsomask/Документы/JasonDipsomask_bot/jdvenv/baskets-in-agreement/xlsx/' + filename
            agreementuser = (filename.replace(".xlsx", "")).replace("user-basket-", "")
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            markup.add(btnToMain)
            aplyOrNot(agreementuser, False)
            jd.send_message(message.chat.id, text=f"Заявка {filename}, клиента https://t.me/{agreementuser} " +
                            "отклонена и удалена.", reply_markup=markup)
            
        elif(message.text == "Скачать общую кокрзину"):
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            markup.add(btnToMain)
            with open(makeXLSXourBasket(), 'rb') as document:
                jd.send_document(message.chat.id, document, caption="Общая корзина", reply_markup=markup)

        elif (message.text == "Создать оповещение"):
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            markup.add(btnToMain)
            jd.send_message(message.chat.id, text="Напишите оповещение для пользователей, ОБЯЗАННО НАЧИНАТЬСЯ СО ЗНАКА /", reply_markup=markup)

        elif message.text.startswith("/") and (not message.text.startswith("/start")):
            notification = message.text.replace("/", "")
            with open(pathes.all_users_chatid, 'r', encoding='utf-8') as file:
                chats = json.load(file)
            for blok in chats:
                jd.send_message(blok["chatid"], text="Объявление: " + f"Здравствуйте {blok["username"]}, " + notification)
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            btn1 = types.KeyboardButton("Показать заявки")
            btn2 = types.KeyboardButton("Скачать общую кокрзину")
            btn3 = types.KeyboardButton("Создать оповещение")
            markup.add(btn1, btn2, btn3)
            jd.send_message(message.chat.id, text="Вы на главной.", reply_markup=markup)

        else:
            jd.send_message(message.chat.id, text="Такая команда не обрабатывается.")




if __name__ == "__main__":
    #bot_polling()
    jd.polling(none_stop=True)
